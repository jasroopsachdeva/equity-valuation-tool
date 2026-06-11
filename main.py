import yfinance as yf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def get_company_data(ticker):
    stock = yf.Ticker(ticker)
    info = stock.info
    
    print("=" * 50)
    print(f"Company: {info.get('longName', 'N/A')}")
    print(f"Sector: {info.get('sector', 'N/A')}")
    print(f"Industry: {info.get('industry', 'N/A')}")
    print("=" * 50)
    
    print("\n--- KEY FINANCIALS ---")
    
    revenue = info.get('totalRevenue', 0)
    ebitda = info.get('ebitda', 0)
    net_income = info.get('netIncomeToCommon', 0)
    
    print(f"Revenue:        ₹{revenue:,.0f}")

    if ebitda == 0 or ebitda is None:
        print("EBITDA:         Not applicable (financial sector)")
        print("                Use P/E and P/B ratios instead")
    else:
        print(f"EBITDA:         ₹{ebitda:,.0f}")

    print(f"Net Income:     ₹{net_income:,.0f}")
    
    print("\n--- MARKET DATA ---")
    
    price = info.get('currentPrice', 0)
    market_cap = info.get('marketCap', 0)
    shares = info.get('sharesOutstanding', 0)
    
    print(f"Current Price:  ₹{price:,.2f}")
    print(f"Market Cap:     ₹{market_cap:,.0f}")
    print(f"Shares Out:     {shares:,.0f}")
    
    return info
def calculate_ratios(info):
    print("\n--- FINANCIAL RATIOS ---")
    
    # Price ratios
    price = info.get('currentPrice', 0)
    eps = info.get('trailingEps', 0)
    book_value = info.get('bookValue', 0)
    
    pe_ratio = price / eps if eps and eps != 0 else None
    pb_ratio = price / book_value if book_value and book_value != 0 else None
    
    # Profitability ratios
    revenue = info.get('totalRevenue', 0)
    ebitda = info.get('ebitda', 0)
    net_income = info.get('netIncomeToCommon', 0)
    gross_profit = info.get('grossProfits', 0)
    
    ebitda_margin = ebitda / revenue if revenue and ebitda else None
    net_margin = net_income / revenue if revenue and net_income else None
    gross_margin = gross_profit / revenue if revenue and gross_profit else None
    
    # Return ratios
    roe = info.get('returnOnEquity', None)
    roa = info.get('returnOnAssets', None)
    
    # Leverage ratio
    total_debt = info.get('totalDebt', 0)
    total_equity = (info.get('totalStockholderEquity') or 
                info.get('stockholdersEquity') or 
                info.get('bookValue', 0) * info.get('sharesOutstanding', 0))
    debt_equity = total_debt / total_equity if total_equity and total_equity != 0 else None
    
    # EV/EBITDA
    enterprise_value = info.get('enterpriseValue', 0)
    ev_ebitda = enterprise_value / ebitda if ebitda and ebitda != 0 else None
    
    # Print everything
    def fmt(value, suffix="x", is_pct=False):
        if value is None:
            return "N/A"
        if is_pct:
            return f"{value * 100:.1f}%"
        return f"{value:.1f}{suffix}"
    
    print(f"P/E Ratio:          {fmt(pe_ratio)}")
    print(f"P/B Ratio:          {fmt(pb_ratio)}")
    print(f"EV/EBITDA:          {fmt(ev_ebitda)}")
    print(f"EBITDA Margin:      {fmt(ebitda_margin, is_pct=True)}")
    print(f"Net Profit Margin:  {fmt(net_margin, is_pct=True)}")
    print(f"Gross Margin:       {fmt(gross_margin, is_pct=True)}")
    print(f"Return on Equity:   {fmt(roe, is_pct=True)}")
    print(f"Debt/Equity:        {fmt(debt_equity)}")
    
    return {
        'pe': pe_ratio, 'pb': pb_ratio, 'ev_ebitda': ev_ebitda,
        'ebitda_margin': ebitda_margin, 'net_margin': net_margin,
        'gross_margin': gross_margin, 'roe': roe, 'debt_equity': debt_equity
    }
def dcf_valuation(info):
    print("\n--- DCF VALUATION ---")
    
    # Get base financial data
    revenue = info.get('totalRevenue', 0)
    ebitda = info.get('ebitda', 0)
    net_income = info.get('netIncomeToCommon', 0)
    shares = info.get('sharesOutstanding', 0)
    total_debt = info.get('totalDebt', 0)
    cash = info.get('totalCash', 0)
    net_debt = total_debt - cash
    
    if revenue == 0 or ebitda == 0:
        print("Insufficient data for DCF valuation")
        return None
    
    # Get assumptions from user
    print("\nEnter DCF assumptions:")
    revenue_growth = float(input("Revenue growth rate (e.g. 0.08 for 8%): "))
    ebitda_margin = float(input("EBITDA margin (e.g. 0.20 for 20%): "))
    wacc = float(input("WACC (e.g. 0.12 for 12%): "))
    terminal_growth = float(input("Terminal growth rate (e.g. 0.04 for 4%): "))
    tax_rate = float(input("Tax rate (e.g. 0.25 for 25%): "))
    capex_pct = float(input("Capex as % of revenue (e.g. 0.03 for 3%): "))
    
    # Project FCF for 10 years
    print("\n--- PROJECTED FREE CASH FLOWS ---")
    print(f"{'Year':<8}{'Revenue':>15}{'EBITDA':>15}{'FCF':>15}{'PV of FCF':>15}")
    print("-" * 58)
    
    projected_revenues = []
    projected_fcfs = []
    pv_fcfs = []
    
    for year in range(1, 11):
        proj_revenue = revenue * (1 + revenue_growth) ** year
        proj_ebitda = proj_revenue * ebitda_margin
        proj_ebit = proj_ebitda * 0.8
        proj_tax = proj_ebit * tax_rate
        proj_capex = proj_revenue * capex_pct
        proj_fcf = proj_ebitda - proj_tax - proj_capex
        
        discount_factor = 1 / (1 + wacc) ** year
        pv_fcf = proj_fcf * discount_factor
        
        projected_revenues.append(proj_revenue)
        projected_fcfs.append(proj_fcf)
        pv_fcfs.append(pv_fcf)
        
        print(f"{year:<8}"
              f"{proj_revenue/1e9:>14.1f}B"
              f"{proj_ebitda/1e9:>14.1f}B"
              f"{proj_fcf/1e9:>14.1f}B"
              f"{pv_fcf/1e9:>14.1f}B")
    
    # Terminal value
    final_fcf = projected_fcfs[-1]
    terminal_value = final_fcf * (1 + terminal_growth) / (wacc - terminal_growth)
    pv_terminal = terminal_value / (1 + wacc) ** 10
    
    # Enterprise and equity value
    sum_pv_fcfs = sum(pv_fcfs)
    enterprise_value = sum_pv_fcfs + pv_terminal
    equity_value = enterprise_value - net_debt
    intrinsic_price = equity_value / shares if shares > 0 else 0
    current_price = info.get('currentPrice', 0)
    
    # Results
    print("\n--- DCF RESULTS ---")
    print(f"PV of FCFs:           ₹{sum_pv_fcfs/1e9:.1f}B  "
          f"({sum_pv_fcfs/enterprise_value*100:.1f}% of EV)")
    print(f"PV of Terminal Value: ₹{pv_terminal/1e9:.1f}B  "
          f"({pv_terminal/enterprise_value*100:.1f}% of EV)")
    print(f"Enterprise Value:     ₹{enterprise_value/1e9:.1f}B")
    print(f"Less: Net Debt:       ₹{net_debt/1e9:.1f}B")
    print(f"Equity Value:         ₹{equity_value/1e9:.1f}B")
    print(f"Intrinsic Price:      ₹{intrinsic_price:,.2f}")
    print(f"Current Price:        ₹{current_price:,.2f}")
    
    upside = (intrinsic_price - current_price) / current_price * 100
    
    if upside > 15:
        recommendation = "BUY"
    elif upside < -15:
        recommendation = "SELL"
    else:
        recommendation = "HOLD"
    
    print(f"Upside/Downside:      {upside:.1f}%")
    print(f"Recommendation:       {recommendation}")
    
    return {
        'intrinsic_price': intrinsic_price,
        'current_price': current_price,
        'upside': upside,
        'recommendation': recommendation,
        'enterprise_value': enterprise_value
    }
def sensitivity_table(info, base_wacc, base_growth):
    print("\n--- SENSITIVITY TABLE (Intrinsic Price ₹) ---")
    print("WACC → / Growth ↓")
    
    wacc_range = [base_wacc - 0.02, base_wacc - 0.01, 
                  base_wacc, base_wacc + 0.01, base_wacc + 0.02]
    growth_range = [base_growth - 0.02, base_growth - 0.01,
                    base_growth, base_growth + 0.01, base_growth + 0.02]
    
    revenue = info.get('totalRevenue', 0)
    ebitda = info.get('ebitda', 0)
    shares = info.get('sharesOutstanding', 0)
    total_debt = info.get('totalDebt', 0)
    cash = info.get('totalCash', 0)
    net_debt = total_debt - cash
    ebitda_margin = ebitda / revenue if revenue else 0
    
    header = f"{'':>8}"
    for w in wacc_range:
        header += f"  {w*100:.1f}%"
    print(header)
    print("-" * 48)
    
    for g in growth_range:
        row = f"{g*100:.1f}%   "
        for w in wacc_range:
            fcfs = []
            for year in range(1, 11):
                proj_rev = revenue * (1 + 0.08) ** year
                proj_fcf = proj_rev * ebitda_margin * 0.75
                pv = proj_fcf / (1 + w) ** year
                fcfs.append(pv)
            terminal = fcfs[-1] * (1 + g) / (w - g) if w > g else 0
            pv_terminal = terminal / (1 + w) ** 10
            ev = sum(fcfs) + pv_terminal
            eq = ev - net_debt
            price = eq / shares if shares else 0
            row += f"  {price:>6.0f}"
        print(row)
def comparable_analysis(main_ticker, main_info):
    print("\n--- COMPARABLE COMPANY ANALYSIS ---")
    
    print("\nEnter peer company tickers for comparison.")
    print("Press Enter without typing anything when done.")
    print("Example: NESTLEIND.NS, BRITANNIA.NS, DABUR.NS")
    print()
    
    peers = {}
    peers[main_ticker] = main_ticker.replace(".NS", "").replace(".BO", "")
    
    while True:
        peer_input = input("Enter peer ticker (or press Enter to finish): ").strip()
        
        if peer_input == "":
            if len(peers) < 2:
                print("Please enter at least one peer company.")
                continue
            break
        
        peer_input = peer_input.upper()
        if not peer_input.endswith(".NS") and not peer_input.endswith(".BO"):
            peer_input = peer_input + ".NS"
        
        short_name = peer_input.replace(".NS", "").replace(".BO", "")
        peers[peer_input] = short_name
        print(f"Added: {short_name}")
    
    results = []
    print("\nFetching data for all companies...")
    
    for ticker, name in peers.items():
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            
            price = info.get('currentPrice', 0)
            revenue = info.get('totalRevenue', 0)
            ebitda = info.get('ebitda', 0)
            net_income = info.get('netIncomeToCommon', 0)
            market_cap = info.get('marketCap', 0)
            enterprise_value = info.get('enterpriseValue', 0)
            eps = info.get('trailingEps', 0)
            
            pe = price / eps if eps and eps != 0 else None
            ev_ebitda = enterprise_value / ebitda if ebitda and ebitda != 0 else None
            ebitda_margin = ebitda / revenue if revenue and ebitda else None
            net_margin = net_income / revenue if revenue and net_income else None
            
            results.append({
                'name': name,
                'ticker': ticker,
                'price': price,
                'market_cap': market_cap / 1e9,
                'pe': pe,
                'ev_ebitda': ev_ebitda,
                'ebitda_margin': ebitda_margin,
                'net_margin': net_margin
            })
            
        except Exception as e:
            print(f"Could not fetch data for {name}: {e}")
    
    print(f"\n{'Company':<12}"
          f"{'Mkt Cap(B)':>12}"
          f"{'P/E':>8}"
          f"{'EV/EBITDA':>12}"
          f"{'EBITDA%':>10}"
          f"{'Net%':>8}")
    print("-" * 62)
    
    pe_values = []
    ev_ebitda_values = []
    
    for r in results:
        pe_str = f"{r['pe']:.1f}x" if r['pe'] else "N/A"
        ev_str = f"{r['ev_ebitda']:.1f}x" if r['ev_ebitda'] else "N/A"
        em_str = f"{r['ebitda_margin']*100:.1f}%" if r['ebitda_margin'] else "N/A"
        nm_str = f"{r['net_margin']*100:.1f}%" if r['net_margin'] else "N/A"
        mc_str = f"{r['market_cap']:.0f}"
        marker = " ← target" if r['ticker'] == main_ticker else ""
        
        print(f"{r['name']:<12}"
              f"{mc_str:>12}"
              f"{pe_str:>8}"
              f"{ev_str:>12}"
              f"{em_str:>10}"
              f"{nm_str:>8}"
              f"{marker}")
        
        if r['pe']:
            pe_values.append(r['pe'])
        if r['ev_ebitda']:
            ev_ebitda_values.append(r['ev_ebitda'])
    
    print("-" * 62)
    avg_pe = sum(pe_values) / len(pe_values) if pe_values else 0
    avg_ev_ebitda = sum(ev_ebitda_values) / len(ev_ebitda_values) if ev_ebitda_values else 0
    
    print(f"{'Peer Average':<12}"
          f"{'':>12}"
          f"{avg_pe:.1f}x{'':<4}"
          f"{avg_ev_ebitda:.1f}x")
    
    print("\n--- IMPLIED VALUATION FROM COMPS ---")
    
    main_ebitda = main_info.get('ebitda', 0)
    main_shares = main_info.get('sharesOutstanding', 0)
    main_net_debt = (main_info.get('totalDebt', 0) -
                     main_info.get('totalCash', 0))
    main_eps = main_info.get('trailingEps', 0)
    
    if avg_ev_ebitda and main_ebitda:
        implied_ev = avg_ev_ebitda * main_ebitda
        implied_equity = implied_ev - main_net_debt
        implied_price_ev = implied_equity / main_shares if main_shares else 0
        print(f"Based on EV/EBITDA ({avg_ev_ebitda:.1f}x): "
              f"₹{implied_price_ev:,.0f} per share")
    
    if avg_pe and main_eps:
        implied_price_pe = avg_pe * main_eps
        print(f"Based on P/E ({avg_pe:.1f}x):        "
              f"₹{implied_price_pe:,.0f} per share")
    
    current = main_info.get('currentPrice', 0)
    print(f"Current market price:          ₹{current:,.0f} per share")
    
    return results

def export_to_excel(ticker, info, ratios, dcf, comps):
    filename = f"outputs/{ticker.replace('.NS','')}_valuation_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Color definitions
    navy_fill = PatternFill("solid", fgColor="003087")
    grey_fill = PatternFill("solid", fgColor="F4F6F9")
    green_fill = PatternFill("solid", fgColor="2D7D46")
    red_fill = PatternFill("solid", fgColor="C0392B")
    gold_fill = PatternFill("solid", fgColor="D4AF37")
    
    def style_header(cell, bg_fill=navy_fill):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = bg_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_label(cell):
        cell.font = Font(bold=True, size=10)
        cell.fill = grey_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def style_value(cell):
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal='right', vertical='center')
    
    def add_border(cell):
        thin = Side(style='thin', color="CCCCCC")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    
    # ── SHEET 1: SUMMARY ──────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 22
    
    # Title
    ws1['A1'] = f"Equity Research Report — {info.get('longName', ticker)}"
    ws1['A1'].font = Font(bold=True, size=14, color="003087")
    ws1['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
    ws1['A2'].font = Font(size=10, color="888888")
    
    # Company info section
    ws1['A4'] = "COMPANY OVERVIEW"
    style_header(ws1['A4'])
    ws1.merge_cells('A4:B4')
    
    overview_data = [
        ("Company", info.get('longName', 'N/A')),
        ("Ticker", ticker),
        ("Sector", info.get('sector', 'N/A')),
        ("Industry", info.get('industry', 'N/A')),
        ("Current Price", f"₹{info.get('currentPrice', 0):,.2f}"),
        ("Market Cap", f"₹{info.get('marketCap', 0)/1e9:.1f}B"),
        ("52W High", f"₹{info.get('fiftyTwoWeekHigh', 0):,.2f}"),
        ("52W Low", f"₹{info.get('fiftyTwoWeekLow', 0):,.2f}"),
    ]
    
    for i, (label, value) in enumerate(overview_data):
        row = i + 5
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
        style_label(ws1.cell(row=row, column=1))
        style_value(ws1.cell(row=row, column=2))
        add_border(ws1.cell(row=row, column=1))
        add_border(ws1.cell(row=row, column=2))
    
    # Key ratios section
    ws1['D4'] = "KEY RATIOS"
    style_header(ws1['D4'])
    
    ratio_data = [
        ("P/E Ratio", f"{ratios.get('pe', 0):.1f}x" if ratios.get('pe') else "N/A"),
        ("EV/EBITDA", f"{ratios.get('ev_ebitda', 0):.1f}x" if ratios.get('ev_ebitda') else "N/A"),
        ("EBITDA Margin", f"{ratios.get('ebitda_margin', 0)*100:.1f}%" if ratios.get('ebitda_margin') else "N/A"),
        ("Net Margin", f"{ratios.get('net_margin', 0)*100:.1f}%" if ratios.get('net_margin') else "N/A"),
        ("Gross Margin", f"{ratios.get('gross_margin', 0)*100:.1f}%" if ratios.get('gross_margin') else "N/A"),
        ("Return on Equity", f"{ratios.get('roe', 0)*100:.1f}%" if ratios.get('roe') else "N/A"),
        ("Debt/Equity", f"{ratios.get('debt_equity', 0):.1f}x" if ratios.get('debt_equity') else "0.0x"),
        ("P/B Ratio", f"{ratios.get('pb', 0):.1f}x" if ratios.get('pb') else "N/A"),
    ]
    
    for i, (label, value) in enumerate(ratio_data):
        row = i + 5
        ws1.cell(row=row, column=4, value=label)
        ws1.cell(row=row, column=5, value=value)
        style_label(ws1.cell(row=row, column=4))
        style_value(ws1.cell(row=row, column=5))
        add_border(ws1.cell(row=row, column=4))
        add_border(ws1.cell(row=row, column=5))
    
    # DCF result box
    if dcf:
        ws1['A14'] = "DCF VALUATION RESULT"
        style_header(ws1['A14'])
        ws1.merge_cells('A14:B14')
        
        dcf_data = [
            ("Intrinsic Price", f"₹{dcf['intrinsic_price']:,.2f}"),
            ("Current Price", f"₹{dcf['current_price']:,.2f}"),
            ("Upside / Downside", f"{dcf['upside']:.1f}%"),
            ("Recommendation", dcf['recommendation']),
        ]
        
        for i, (label, value) in enumerate(dcf_data):
            row = i + 15
            ws1.cell(row=row, column=1, value=label)
            ws1.cell(row=row, column=2, value=value)
            style_label(ws1.cell(row=row, column=1))
            style_value(ws1.cell(row=row, column=2))
            add_border(ws1.cell(row=row, column=1))
            add_border(ws1.cell(row=row, column=2))
        
        rec_cell = ws1.cell(row=18, column=2)
        if dcf['recommendation'] == "BUY":
            rec_cell.fill = green_fill
            rec_cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif dcf['recommendation'] == "SELL":
            rec_cell.fill = red_fill
            rec_cell.font = Font(bold=True, color="FFFFFF", size=10)
        else:
            rec_cell.fill = gold_fill
            rec_cell.font = Font(bold=True, color="FFFFFF", size=10)
    
    # ── SHEET 2: COMPS TABLE ──────────────────────
    ws2 = wb.create_sheet("Comps Table")
    ws2.column_dimensions['A'].width = 16
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 16
    
    ws2['A1'] = "COMPARABLE COMPANY ANALYSIS"
    ws2['A1'].font = Font(bold=True, size=13, color="003087")
    
    headers = ["Company", "Mkt Cap (₹B)", "P/E", "EV/EBITDA", "EBITDA Margin", "Net Margin"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        style_header(cell)
        add_border(cell)
    
    for row_idx, comp in enumerate(comps, 4):
        data_row = [
            comp['name'],
            f"{comp['market_cap']:.0f}",
            f"{comp['pe']:.1f}x" if comp['pe'] else "N/A",
            f"{comp['ev_ebitda']:.1f}x" if comp['ev_ebitda'] else "N/A",
            f"{comp['ebitda_margin']*100:.1f}%" if comp['ebitda_margin'] else "N/A",
            f"{comp['net_margin']*100:.1f}%" if comp['net_margin'] else "N/A",
        ]
        for col_idx, value in enumerate(data_row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = grey_fill
            style_value(cell)
            add_border(cell)
    
    wb.save(filename)
    print(f"\nReport saved: {filename}")
    return filename

ticker = input("Enter stock ticker (e.g. HINDUNILVR.NS): ")
data = get_company_data(ticker)
ratios = calculate_ratios(data)
dcf = dcf_valuation(data)
if dcf:
    wacc = float(input("\nEnter WACC for sensitivity table: "))
    growth = float(input("Enter terminal growth for sensitivity table: "))
    sensitivity_table(data, wacc, growth)
comps = comparable_analysis(ticker, data)
export_to_excel(ticker, data, ratios, dcf, comps)