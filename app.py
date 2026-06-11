import streamlit as st
import yfinance as yf
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io

st.set_page_config(
    page_title="Equity Valuation Tool",
    page_icon="📊",
    layout="wide"
)

st.title("Equity Valuation and Financial Analysis Tool")
st.divider()

col_input, col_button = st.columns([4, 1])

with col_input:
    ticker = st.text_input(
        "Enter Stock Ticker",
        placeholder="e.g. HINDUNILVR.NS, TCS.NS, AAPL"
    )

with col_button:
    st.write("")
    st.write("")
    analyse = st.button("Analyse", type="primary", use_container_width=True)

if analyse and ticker:
    with st.spinner("Fetching data..."):
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            name = info.get('longName', ticker)
            price = info.get('currentPrice', 0)
            market_cap = info.get('marketCap', 0)
            sector = info.get('sector', 'N/A')
            currency = info.get('currency', 'INR')
            currency_symbol = '₹' if currency == 'INR' else '$' if currency == 'USD' else currency + ' '

            if price == 0:
                st.error("No data found. Check the ticker symbol and try again.")
                st.stop()

            st.session_state['info'] = info
            st.session_state['ticker'] = ticker
            st.session_state['name'] = name
            st.session_state['price'] = price
            st.session_state['market_cap'] = market_cap
            st.session_state['sector'] = sector
            st.session_state['currency_symbol'] = currency_symbol
            st.session_state['loaded'] = True

        except Exception as e:
            st.error(f"Could not fetch data: {e}")

if st.session_state.get('loaded'):
    info = st.session_state['info']
    ticker = st.session_state['ticker']
    name = st.session_state['name']
    price = st.session_state['price']
    market_cap = st.session_state['market_cap']
    sector = st.session_state['sector']
    currency_symbol = st.session_state['currency_symbol']

    st.success(f"Loaded: {name}")
    st.divider()

    # ── SECTION 1: COMPANY OVERVIEW ──────────────
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Current Price", f"{currency_symbol}{price:,.2f}")
    with col2:
        st.metric("Market Cap", f"{currency_symbol}{market_cap/1e9:.1f}B")
    with col3:
        st.metric("Sector", sector)
    with col4:
        high = info.get('fiftyTwoWeekHigh', 0)
        low = info.get('fiftyTwoWeekLow', 0)
        st.metric("52W Range", f"{currency_symbol}{low:,.0f} – {currency_symbol}{high:,.0f}")

    # ── SECTION 2: FINANCIAL RATIOS ──────────────
    st.divider()
    st.subheader("Financial Ratios")

    revenue = info.get('totalRevenue', 0)
    ebitda = info.get('ebitda', 0)
    net_income = info.get('netIncomeToCommon', 0)
    gross_profit = info.get('grossProfits', 0)
    eps = info.get('trailingEps', 0)
    book_value = info.get('bookValue', 0)
    enterprise_value = info.get('enterpriseValue', 0)
    total_debt = info.get('totalDebt', 0)
    cash = info.get('totalCash', 0)
    net_debt = total_debt - cash
    shares = info.get('sharesOutstanding', 0)
    total_equity = (info.get('totalStockholderEquity') or
                    info.get('stockholdersEquity') or
                    info.get('bookValue', 0) * info.get('sharesOutstanding', 0))
    roe = info.get('returnOnEquity', None)

    pe = price / eps if eps and eps != 0 else None
    pb = price / book_value if book_value and book_value != 0 else None
    ev_ebitda = enterprise_value / ebitda if ebitda and ebitda != 0 else None
    ebitda_margin = ebitda / revenue if revenue and ebitda else None
    net_margin = net_income / revenue if revenue and net_income else None
    gross_margin = gross_profit / revenue if revenue and gross_profit else None
    debt_equity = total_debt / total_equity if total_equity and total_equity != 0 else 0

    def fmt_ratio(val, suffix="x", is_pct=False):
        if val is None:
            return "N/A"
        if is_pct:
            return f"{val*100:.1f}%"
        return f"{val:.1f}{suffix}"

    ratio_data = {
        "Metric": [
            "P/E Ratio", "P/B Ratio", "EV/EBITDA",
            "EBITDA Margin", "Net Profit Margin",
            "Gross Margin", "Return on Equity", "Debt/Equity"
        ],
        "Value": [
            fmt_ratio(pe), fmt_ratio(pb), fmt_ratio(ev_ebitda),
            fmt_ratio(ebitda_margin, is_pct=True),
            fmt_ratio(net_margin, is_pct=True),
            fmt_ratio(gross_margin, is_pct=True),
            fmt_ratio(roe, is_pct=True),
            "0.0x (debt-free)" if debt_equity == 0 else fmt_ratio(debt_equity)
        ],
        "What It Means": [
            "How much investors pay per ₹1 of earnings",
            "Price vs accounting book value",
            "Most common IB valuation multiple",
            "% of revenue that becomes EBITDA",
            "% of revenue that becomes net profit",
            "% of revenue after cost of goods sold",
            "Profit generated per ₹1 of equity",
            "Debt relative to equity — lower is safer"
        ]
    }

    df_ratios = pd.DataFrame(ratio_data)
    st.dataframe(df_ratios, use_container_width=True, hide_index=True)

    ratios = {
        'pe': pe, 'pb': pb, 'ev_ebitda': ev_ebitda,
        'ebitda_margin': ebitda_margin, 'net_margin': net_margin,
        'gross_margin': gross_margin, 'roe': roe,
        'debt_equity': debt_equity
    }

    # ── SECTION 3: DCF VALUATION ──────────────────
    st.divider()
    st.subheader("DCF Valuation")

    with st.expander("Set DCF Assumptions", expanded=True):
        dcf_col1, dcf_col2, dcf_col3 = st.columns(3)
        with dcf_col1:
            revenue_growth = st.slider("Revenue Growth Rate", 1, 20, 10, 1,
                                       help="Expected annual revenue growth %") / 100
            wacc = st.slider("WACC", 6, 20, 12, 1,
                             help="Weighted average cost of capital %") / 100
        with dcf_col2:
            ebitda_margin_input = st.slider("EBITDA Margin", 5, 50, 22, 1,
                                            help="Projected EBITDA as % of revenue") / 100
            terminal_growth = st.slider("Terminal Growth Rate", 1, 8, 5, 1,
                                        help="Long run perpetual growth rate %") / 100
        with dcf_col3:
            tax_rate = st.slider("Tax Rate", 10, 35, 25, 1,
                                 help="Effective corporate tax rate %") / 100
            capex_pct = st.slider("Capex % of Revenue", 1, 10, 3, 1,
                                  help="Capital expenditure as % of revenue") / 100

    if revenue == 0 or ebitda == 0:
        st.warning("Insufficient financial data for DCF valuation.")
    else:
        projected_fcfs = []
        pv_fcfs = []
        years_list = []
        rev_list = []
        ebitda_list = []

        for year in range(1, 11):
            proj_rev = revenue * (1 + revenue_growth) ** year
            proj_ebitda = proj_rev * ebitda_margin_input
            proj_ebit = proj_ebitda * 0.8
            proj_tax = proj_ebit * tax_rate
            proj_capex = proj_rev * capex_pct
            proj_fcf = proj_ebitda - proj_tax - proj_capex
            pv_fcf = proj_fcf / (1 + wacc) ** year

            years_list.append(f"Year {year}")
            rev_list.append(round(proj_rev / 1e9, 1))
            ebitda_list.append(round(proj_ebitda / 1e9, 1))
            projected_fcfs.append(proj_fcf)
            pv_fcfs.append(pv_fcf)

        terminal_value = projected_fcfs[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
        pv_terminal = terminal_value / (1 + wacc) ** 10
        sum_pv_fcfs = sum(pv_fcfs)
        ev_dcf = sum_pv_fcfs + pv_terminal
        equity_value = ev_dcf - net_debt
        intrinsic_price = equity_value / shares if shares > 0 else 0
        upside = (intrinsic_price - price) / price * 100 if price > 0 else 0

        if upside > 15:
            recommendation = "BUY"
        elif upside < -15:
            recommendation = "SELL"
        else:
            recommendation = "HOLD"

        dcf_r1, dcf_r2, dcf_r3, dcf_r4 = st.columns(4)
        with dcf_r1:
            st.metric("Intrinsic Price",
                      f"{currency_symbol}{intrinsic_price:,.2f}",
                      f"{upside:+.1f}% vs market")
        with dcf_r2:
            st.metric("Enterprise Value", f"{currency_symbol}{ev_dcf/1e9:.1f}B")
        with dcf_r3:
            st.metric("PV of FCFs",
                      f"{currency_symbol}{sum_pv_fcfs/1e9:.1f}B",
                      f"{sum_pv_fcfs/ev_dcf*100:.1f}% of EV")
        with dcf_r4:
            st.metric("PV of Terminal Value",
                      f"{currency_symbol}{pv_terminal/1e9:.1f}B",
                      f"{pv_terminal/ev_dcf*100:.1f}% of EV")

        if recommendation == "BUY":
            st.success(f"Recommendation: {recommendation} — Intrinsic value {upside:.1f}% above market price")
        elif recommendation == "SELL":
            st.error(f"Recommendation: {recommendation} — Intrinsic value {abs(upside):.1f}% below market price")
        else:
            st.warning(f"Recommendation: {recommendation} — Trading close to intrinsic value ({upside:.1f}%)")

        df_proj = pd.DataFrame({
            "Year": years_list,
            f"Revenue ({currency_symbol}B)": rev_list,
            f"EBITDA ({currency_symbol}B)": ebitda_list,
            f"FCF ({currency_symbol}B)": [round(f/1e9, 1) for f in projected_fcfs],
            f"PV of FCF ({currency_symbol}B)": [round(p/1e9, 1) for p in pv_fcfs]
        })
        st.dataframe(df_proj, use_container_width=True, hide_index=True)

        dcf_result = {
            'intrinsic_price': intrinsic_price,
            'current_price': price,
            'upside': upside,
            'recommendation': recommendation,
            'enterprise_value': ev_dcf
        }

    # ── SECTION 4: COMPS TABLE ────────────────────
    st.divider()
    st.subheader("Comparable Company Analysis")

    peer_input = st.text_input(
        "Enter peer tickers separated by commas",
        placeholder="e.g. NESTLEIND.NS, BRITANNIA.NS, DABUR.NS",
        key="peer_input"
    )

    run_comps = st.button("Run Comps", type="secondary")

    if run_comps and peer_input:
        peer_tickers = [p.strip().upper() for p in peer_input.split(",")]
        all_tickers = [ticker] + peer_tickers
        comp_data = []

        with st.spinner("Fetching peer data..."):
            for t in all_tickers:
                try:
                    s = yf.Ticker(t)
                    i = s.info
                    t_price = i.get('currentPrice', 0)
                    t_rev = i.get('totalRevenue', 0)
                    t_ebitda = i.get('ebitda', 0)
                    t_ni = i.get('netIncomeToCommon', 0)
                    t_mc = i.get('marketCap', 0)
                    t_ev = i.get('enterpriseValue', 0)
                    t_eps = i.get('trailingEps', 0)

                    t_pe = t_price / t_eps if t_eps and t_eps != 0 else None
                    t_ev_ebitda = t_ev / t_ebitda if t_ebitda and t_ebitda != 0 else None
                    t_em = t_ebitda / t_rev if t_rev and t_ebitda else None
                    t_nm = t_ni / t_rev if t_rev and t_ni else None

                    comp_data.append({
                        "Company": t.replace(".NS","").replace(".BO",""),
                        f"Mkt Cap ({currency_symbol}B)": f"{t_mc/1e9:.0f}",
                        "P/E": fmt_ratio(t_pe),
                        "EV/EBITDA": fmt_ratio(t_ev_ebitda),
                        "EBITDA Margin": fmt_ratio(t_em, is_pct=True),
                        "Net Margin": fmt_ratio(t_nm, is_pct=True),
                        "": "← target" if t == ticker else ""
                    })
                except:
                    pass

        st.session_state['comp_data'] = comp_data

    if st.session_state.get('comp_data'):
        comp_data = st.session_state['comp_data']
        df_comps = pd.DataFrame(comp_data)
        st.dataframe(df_comps, use_container_width=True, hide_index=True)

        pe_vals = []
        ev_vals = []
        for r in comp_data:
            try:
                if r["P/E"] != "N/A":
                    pe_vals.append(float(r["P/E"].replace("x","")))
                if r["EV/EBITDA"] != "N/A":
                    ev_vals.append(float(r["EV/EBITDA"].replace("x","")))
            except:
                pass

        if pe_vals and ev_vals:
            avg_pe = sum(pe_vals) / len(pe_vals)
            avg_ev = sum(ev_vals) / len(ev_vals)

            st.markdown(f"**Peer Average P/E:** {avg_pe:.1f}x  |  **Peer Average EV/EBITDA:** {avg_ev:.1f}x")

            imp_col1, imp_col2, imp_col3 = st.columns(3)
            with imp_col1:
                if eps:
                    imp_pe = avg_pe * eps
                    st.metric("Implied Price (P/E)", f"{currency_symbol}{imp_pe:,.0f}")
            with imp_col2:
                if ebitda and shares:
                    imp_ev2 = avg_ev * ebitda
                    imp_eq = imp_ev2 - net_debt
                    imp_ev_price = imp_eq / shares
                    st.metric("Implied Price (EV/EBITDA)", f"{currency_symbol}{imp_ev_price:,.0f}")
            with imp_col3:
                st.metric("Current Market Price", f"{currency_symbol}{price:,.0f}")

    # ── EXCEL DOWNLOAD ────────────────────────────
    st.divider()
    st.subheader("Download Report")

    if st.button("Generate Excel Report", type="primary"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        navy_fill = PatternFill("solid", fgColor="003087")
        grey_fill = PatternFill("solid", fgColor="F4F6F9")
        thin = Side(style='thin', color="CCCCCC")
        bdr = Border(top=thin, bottom=thin, left=thin, right=thin)

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 18

        ws['A1'] = f"Equity Research Report — {name}"
        ws['A1'].font = Font(bold=True, size=14, color="003087")
        ws['A2'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
        ws['A2'].font = Font(size=10, color="888888")

        ws['A4'] = "COMPANY OVERVIEW"
        ws['A4'].font = Font(bold=True, color="FFFFFF", size=11)
        ws['A4'].fill = navy_fill
        ws['A4'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A4:B4')

        for idx, (label, val) in enumerate([
            ("Company", name), ("Ticker", ticker),
            ("Sector", sector), ("Price", f"{currency_symbol}{price:,.2f}"),
            ("Market Cap", f"{currency_symbol}{market_cap/1e9:.1f}B"),
        ]):
            row = idx + 5
            c1 = ws.cell(row=row, column=1, value=label)
            c2 = ws.cell(row=row, column=2, value=val)
            c1.font = Font(bold=True, size=10)
            c1.fill = grey_fill
            c2.font = Font(size=10)
            c2.alignment = Alignment(horizontal='right')
            c1.border = bdr
            c2.border = bdr

        ws['D4'] = "KEY RATIOS"
        ws['D4'].font = Font(bold=True, color="FFFFFF", size=11)
        ws['D4'].fill = navy_fill
        ws['D4'].alignment = Alignment(horizontal='center')
        ws.merge_cells('D4:E4')

        for idx, (label, val) in enumerate([
            ("P/E Ratio", fmt_ratio(pe)),
            ("EV/EBITDA", fmt_ratio(ev_ebitda)),
            ("EBITDA Margin", fmt_ratio(ebitda_margin, is_pct=True)),
            ("Net Margin", fmt_ratio(net_margin, is_pct=True)),
            ("Return on Equity", fmt_ratio(roe, is_pct=True)),
        ]):
            row = idx + 5
            c1 = ws.cell(row=row, column=4, value=label)
            c2 = ws.cell(row=row, column=5, value=val)
            c1.font = Font(bold=True, size=10)
            c1.fill = grey_fill
            c2.font = Font(size=10)
            c2.alignment = Alignment(horizontal='right')
            c1.border = bdr
            c2.border = bdr

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="⬇ Download Excel Report",
            data=buffer,
            file_name=f"{ticker.replace('.NS','')}_valuation_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )